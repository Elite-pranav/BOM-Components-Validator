"""
Excel BOM extractor.

Reads PUMP BOM Excel spreadsheets (.XLSX) exported from SAP and extracts
every line item as a clean dict.

Column discovery
----------------
Instead of hardcoding column indices, the extractor reads the first row as a
header and matches each header cell against known synonym patterns. This makes
it robust to slight naming variations across SAP export configurations.

The four core columns are always required:
  item_number      — "Item", "Item No.", "Itm", "SL No", etc.
  component_number — "Component", "Component number", "Object", "Material No"
  description      — "Description", "Object description", "Part Description"
  quantity         — "Quantity", "Qty", "Comp. Qty", "Comp. Qty (CUn)"

Optional columns (left as None if absent):
  unit             — "Unit", "UoM", "Base Unit of Measure"
  text1            — "Text Line 1", "Item Text Line 1", "Text 1"
  text2            — "Text Line 2", "Item Text Line 2", "Text 2"
  sort_string      — "Sort String", "Sort"

Output shape
------------
[
  {
    "item_number":       "0010",
    "component_number":  "8263538",
    "description":       "STRAINER 2 5638 4900 0501 SS304",
    "quantity":          1,
    "unit":              "PC",
    "text1":             "G.A.DRG.NO.:813351387-40 GA.",
    "text2":             "C.S.DRG.NO.:813351387-40 CS.",
    "sort_string":       "PL BOWL"
  },
  ...
]
"""

import re
from pathlib import Path

import openpyxl

from backend.extractors.base import BaseExtractor

# ── Column synonym patterns ───────────────────────────────────────────────────
# Each entry: (output_field, regex_pattern, required)
_COLUMN_PATTERNS = [
    ("item_number",       re.compile(r"item|itm|sl\.?\s*no|sr\.?\s*no",           re.I), True),
    ("component_number",  re.compile(r"component|material\s*no|object\s*(no|num)", re.I), True),
    ("description",       re.compile(r"desc",                                       re.I), True),
    ("quantity",          re.compile(r"qty|quantity|comp\.?\s*qty",                re.I), True),
    ("unit",              re.compile(r"unit|uom",                                   re.I), False),
    ("text1",             re.compile(r"text.*(line\s*1|1$)|item\s*text\s*1",       re.I), False),
    ("text2",             re.compile(r"text.*(line\s*2|2$)|item\s*text\s*2",       re.I), False),
    ("sort_string",       re.compile(r"sort",                                       re.I), False),
]


class BOMExtractor(BaseExtractor):
    """Extracts all line items from a SAP-exported BOM Excel file."""

    def extract(self) -> list:
        xlsx_file = self._find_xlsx()
        if not xlsx_file:
            self.logger.error(f"No BOM XLSX found in {self.raw_folder}")
            return []

        col_map, data_rows = self._read_excel(xlsx_file)
        if col_map is None:
            return []
        if not data_rows:
            self.logger.warning("No data rows found in BOM Excel")
            return []

        items = [self._parse_row(row, col_map) for row in data_rows]
        self.logger.info(f"Extracted {len(items)} line items from BOM Excel")
        self._save_json(items, self.processed_folder / "bom_data.json")
        return items

    # ── File discovery ────────────────────────────────────────────────────────

    def _find_xlsx(self) -> Path | None:
        for pattern in ("*BOM.XLSX", "*BOM.xlsx", "*bom.xlsx"):
            matches = list(self.raw_folder.glob(pattern))
            if matches:
                return matches[0]
        return None

    # ── Reading ───────────────────────────────────────────────────────────────

    def _read_excel(self, xlsx_path: Path) -> tuple[dict | None, list[list]]:
        """
        Read the Excel file.

        Returns (col_map, data_rows) where:
          col_map   — {field_name: column_index} built from the header row
          data_rows — list of raw row lists (header row excluded)
        """
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        header_row = None
        data_rows  = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row = list(row)
            if i == 0:
                header_row = row
                continue
            if all(v is None for v in row):
                continue
            data_rows.append(row)

        wb.close()

        col_map = self._map_columns(header_row or [])
        if col_map is None:
            self.logger.error(
                f"Could not find required columns in BOM Excel header: {header_row}"
            )
            return None, []

        self.logger.info(f"BOM column map: {col_map}")
        return col_map, data_rows

    def _map_columns(self, header: list) -> dict | None:
        """
        Match header cells against synonym patterns and return a column index map.
        Returns None if any required column cannot be matched.
        """
        col_map: dict[str, int] = {}
        normalised = [str(h).strip() if h is not None else "" for h in header]

        for field, pattern, required in _COLUMN_PATTERNS:
            idx = next(
                (i for i, h in enumerate(normalised) if h and pattern.search(h)),
                None,
            )
            if idx is not None:
                col_map[field] = idx
            elif required:
                return None   # required column missing

        return col_map

    # ── Row parsing ───────────────────────────────────────────────────────────

    def _parse_row(self, row: list, col_map: dict) -> dict:
        def _get(field):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        def _str(val) -> str | None:
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        def _qty(val) -> int | float | None:
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return val
            try:
                f = float(str(val).strip())
                return int(f) if f == int(f) else f
            except ValueError:
                return None

        return {
            "item_number":      _str(_get("item_number")),
            "component_number": _str(_get("component_number")),
            "description":      _str(_get("description")),
            "quantity":         _qty(_get("quantity")),
            "unit":             _str(_get("unit")),
            "text1":            _str(_get("text1")),
            "text2":            _str(_get("text2")),
            "sort_string":      _str(_get("sort_string")),
        }
